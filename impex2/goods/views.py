from django.shortcuts import render
import os
import openpyxl
from decimal import Decimal
from . models import Good
# Create your views here.

# импорт прайса ITP
def export_to_sqlite():
    # Экспорт данных из xlsx в goods
    # 1. Работа c xlsx файлом
    file_to_read = openpyxl.load_workbook('price.xlsx', data_only=True)
    sheet = file_to_read['Sheet']
    # Цикл по строкам

    a = sheet.max_row
    cat1 = cat2 = cat3 = 0
    begin = 17179 # начальная строка импорта
    end = 26453 # конечная строка импорта
    for row in range(begin, end + 1):
        # обработка строки
        article = sheet.cell(row, 4).value
        cat1_name = sheet.cell(row, 1).value
        cat2_name = sheet.cell(row, 2).value
        cat3_name = sheet.cell(row, 3).value
        name = sheet.cell(row, 7).value
        catalog_number = sheet.cell(row, 5).value
        in_stock = Decimal(sheet.cell(row, 8).value) if str(sheet.cell(row, 8).value).isnumeric() else Decimal('0')
        price = Decimal(sheet.cell(row, 9).value) if str(sheet.cell(row, 9).value).isnumeric() else Decimal('0')
        # проверка, что это раздел 1-ого уровня
        if sheet.cell(row, 1).value is not None and sheet.cell(row, 2).value is None:
            # добавление раздела 1-ого уровня в базу
            p = Good(name=cat1_name, article='', catalog_number='', in_stock=0, is_good=False, cat1=0, cat2=0, cat3=0,
                     price=0, supplier='ITP', description='')
            p.save()
            cat1 = p.pk
            cat2 = 0
            cat3 = 0
        # проверка, что это раздел 2-ого уровня
        elif sheet.cell(row, 1).value is not None and sheet.cell(row, 2).value is not None and sheet.cell(row, 3).value is None:
            # добавление раздела 2-ого уровня в базу
            p = Good(name=cat2_name, article='', catalog_number='', in_stock=0, is_good=False, cat1=cat1, cat2=0, cat3=0,
                     price=0, supplier='ITP', description='')
            p.save()
            cat2 = p.pk
            cat3 = 0
        # проверка, что это раздел 3-ого уровня
        elif sheet.cell(row, 1).value is not None and sheet.cell(row, 2).value is not None and sheet.cell(row, 3).value is not None and sheet.cell(row, 7).value is None:
            # добавление раздела 2-ого уровня в базу
            p = Good(name=cat3_name, article='', catalog_number='', in_stock=0, is_good=False, cat1=cat1, cat2=cat2, cat3=0,
                     price=0, supplier='ITP', description='')
            p.save()
            cat3 = p.pk
        elif sheet.cell(row, 7).value is not None:
            # добавление товара в базу
            p = Good(name=name, article=article, catalog_number=catalog_number, in_stock=in_stock, is_good=True, cat1=cat1, cat2=cat2, cat3=cat3,
                     price=price, supplier='ITP', description='')
            p.save()
        print('обработано ' + str(row + 1 - begin) + ' строк из ' + str(end + 1 - begin), end='')

    return None


# импорт прайса VTT
def export_to_sqlite2():
    #'Экспорт данных из xlsx в sqlite
    # 1.Работа c xlsx файлом

    # Читаем файл и лист1 книги excel
    file_to_read = openpyxl.load_workbook('rtp.xlsx', data_only=True)
    sheet = file_to_read['Price']
    # Цикл по строкам начиная с шестой (выше заголовки)

    a = sheet.max_row
    cat1 = cat2 = 0
    for row in range(6, a - 1):
        # обработка строки
        article = sheet.cell(row, 1).value
        cat_name = sheet.cell(row, 2).value
        name = sheet.cell(row, 4).value
        catalog_number = sheet.cell(row, 6).value
        in_stock = Decimal(sheet.cell(row, 8).value) if str(sheet.cell(row, 8).value).isnumeric() else Decimal('0')
        price = Decimal(sheet.cell(row, 9).value) if str(sheet.cell(row, 9).value).isnumeric() else Decimal('0')
        # проверка, что это раздел 1-ого уровня
        if sheet.cell(row, 4).value is None and sheet.cell(row + 1, 4).value is None:
            # добавление раздела 1-ого уровня в базу
            p = Good(name=cat_name, article='', catalog_number='', in_stock=0, is_good=False, cat1=0, cat2=0, cat3=0,
                     price=0, supplier='VTT', description='')
            p.save()
            cat1 = p.pk
            cat2 = 0
        # проверка, что это раздел 2-ого уровня
        elif sheet.cell(row, 4).value is None and sheet.cell(row + 1, 4).value is not None:
            # добавление раздела 2-ого уровня в базу
            p = Good(name=cat_name, article='', catalog_number='', in_stock=0, is_good=False, cat1=cat1, cat2=0, cat3=0,
                     price=0, supplier='VTT', description='')
            p.save()
            cat2 = p.pk
        else:
            # добавление товара в базу
            p = Good(name=name, article=article, catalog_number=catalog_number, in_stock=in_stock, is_good=True, cat1=cat1, cat2=cat2, cat3=0,
                     price=price, supplier='VTT', description='')
            p.save()
        print('обработано ' + str(row) + ' строк из ' + str(a), end='')

    return None